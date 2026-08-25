"""Deterministic, sanitized XLSX payloads generated only for tests."""

from datetime import datetime, timezone
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook

_WORKSHEET_XML = "xl/worksheets/sheet1.xml"
_FIXED_TIMESTAMP = (2026, 8, 20, 0, 0, 0)


def make_xlsx(*, headers: list[object], rows: list[list[object]], malformed_dimension: bool = False) -> bytes:
    """Build a one-sheet workbook, optionally corrupting only its declared range."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Отчёт"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    fixed_time = datetime(2026, 8, 20, tzinfo=timezone.utc)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    payload = stream.getvalue()
    if malformed_dimension:
        payload = _replace_dimension_with_a1(payload)
    return _normalize_zip_metadata(payload)


def make_multisheet_xlsx(sheets: list[tuple[str, list[object], list[list[object]]]]) -> bytes:
    """Build a deterministic workbook with explicitly named worksheets."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, headers, rows in sheets:
        worksheet = workbook.create_sheet(name)
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)
    fixed_time = datetime(2026, 8, 20, tzinfo=timezone.utc)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return _normalize_zip_metadata(stream.getvalue())


def make_real_unitka(*, product_rows=None, tariff_rows=None, fbo_complete=True,
                     product_available_qty=None) -> bytes:
    """Build the sanitized two-sheet shape used by operational Unitka files."""
    workbook = Workbook()
    economics = workbook.active
    economics.title = "Расчёт юнит-экономики"
    for _ in range(8):
        economics.append([])
    product_headers = ["Артикул", "Название товара", "Себестоимость единицы",
                       "Цена поставщика до скидок OZON", "Комиссия OZON %", "Объём товара (л)"]
    if product_available_qty is not None:
        product_headers.append("Доступный остаток")
    economics.append(product_headers)
    for row in product_rows or [["ART-1", "Товар", 100, 1000, "10%", 1]]:
        economics.append([*row, product_available_qty] if product_available_qty is not None else row)
    tariffs = workbook.create_sheet("Логистика с 28 августа 2026г.")
    tariffs.cell(2, 2, "FBO"); tariffs.cell(2, 12, "FBS"); tariffs.cell(2, 23, "Базовый тариф")
    headers = ["Диапазон объёма ОТ", "RANG", "key", "Объём товара", "Кластер поставки",
               "Кластер доставки", "Для товаров до 300 руб.", "Для товаров свыше 300 руб."]
    for start in (2, 12, 23):
        for offset, header in enumerate(headers): tariffs.cell(4, start + offset, header)
    if not fbo_complete:
        tariffs.cell(4, 9).value = None
    rows = tariff_rows or [(0, "0-0,200 л", "Москва", "Москва", 18, 69)]
    for index, (minimum, label, origin, destination, low, high) in enumerate(rows, 5):
        values = [minimum, "", "", label, origin, destination, low, high]
        for offset, value in enumerate(values): tariffs.cell(index, 2 + offset, value)
        fallback_volume = minimum if not fbo_complete else label
        for offset, value in enumerate([minimum + 1000, "", "", fallback_volume, origin, destination, low + 100, high + 100]): tariffs.cell(index, 12 + offset, value)
        for offset, value in enumerate([minimum, "", "", fallback_volume, origin, destination, 5, 15]): tariffs.cell(index, 23 + offset, value)
    stream = BytesIO(); workbook.save(stream); workbook.close()
    return _normalize_zip_metadata(stream.getvalue())


def worksheet_xml(payload: bytes) -> bytes:
    """Expose worksheet XML so tests can independently prove fixture shape."""
    with ZipFile(BytesIO(payload)) as archive:
        return archive.read(_WORKSHEET_XML)


def _replace_dimension_with_a1(payload: bytes) -> bytes:
    source = BytesIO(payload)
    target = BytesIO()
    with ZipFile(source) as archive, ZipFile(target, "w", ZIP_DEFLATED) as rebuilt:
        for name in archive.namelist():
            content = archive.read(name)
            if name == _WORKSHEET_XML:
                start = content.index(b"<dimension ")
                end = content.index(b"/>", start) + 2
                content = content[:start] + b'<dimension ref="A1"/>' + content[end:]
            rebuilt.writestr(name, content)
    return target.getvalue()


def _normalize_zip_metadata(payload: bytes) -> bytes:
    target = BytesIO()
    with ZipFile(BytesIO(payload)) as archive, ZipFile(target, "w", ZIP_DEFLATED) as rebuilt:
        for name in sorted(archive.namelist()):
            info = ZipInfo(name, _FIXED_TIMESTAMP)
            info.compress_type = ZIP_DEFLATED
            info.external_attr = archive.getinfo(name).external_attr
            rebuilt.writestr(info, archive.read(name))
    return target.getvalue()
