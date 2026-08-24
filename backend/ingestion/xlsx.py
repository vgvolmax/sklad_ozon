"""openpyxl worksheet adapter with malformed declared-dimension recovery."""

from dataclasses import dataclass
from typing import BinaryIO

from openpyxl import load_workbook

from backend.domain.contracts import ImportDiagnostic


@dataclass(frozen=True, slots=True)
class WorksheetRow:
    source_row: int
    values: tuple[object, ...]


@dataclass(frozen=True, slots=True)
class WorksheetRows:
    rows: tuple[WorksheetRow, ...]
    diagnostics: tuple[ImportDiagnostic, ...]


def iter_worksheet_rows(stream: BinaryIO, sheet_selector: int | str) -> WorksheetRows:
    workbook = load_workbook(stream, read_only=True, data_only=True)
    worksheet = workbook.worksheets[sheet_selector] if isinstance(sheet_selector, int) else workbook[sheet_selector]
    suspected_truncated_dimension = worksheet.max_row == 1 and worksheet.max_column == 1
    if suspected_truncated_dimension:
        worksheet.reset_dimensions()
    rows = tuple(
        WorksheetRow(source_row=index, values=tuple(cell.value for cell in cells))
        for index, cells in enumerate(worksheet.iter_rows(), start=1)
        if any(cell.value is not None for cell in cells)
    )
    workbook.close()
    diagnostics = ()
    actually_repaired = suspected_truncated_dimension and (
        len(rows) > 1 or (rows and any(value is not None for value in rows[0].values[1:]))
    )
    if actually_repaired:
        diagnostics = (ImportDiagnostic(
            severity="warning", code="WORKSHEET_DIMENSION_REPAIRED",
            message="Worksheet declared a single-cell range but contained populated cells outside A1.",
        ),)
    return WorksheetRows(rows=rows, diagnostics=diagnostics)
