"""Validated tariff workbook importer with signature-based sheet selection."""

from io import BytesIO
from collections import defaultdict
from decimal import Decimal
import re
from openpyxl import load_workbook

from backend.domain.contracts import ImportResult, ReportMeta, TariffRow
from ._common import _diag, parse_decimal
from .normalization import normalize_header, normalize_text

_ALIASES = {
    "кластер отгрузки": "origin", "кластер отправления": "origin", "origin cluster": "origin",
    "кластер доставки": "destination", "destination cluster": "destination",
    "объём от": "min_volume", "объем от": "min_volume", "min volume": "min_volume",
    "объём до": "max_volume", "объем до": "max_volume", "max volume": "max_volume",
    "цена от": "min_price", "min price": "min_price", "цена до": "max_price", "max price": "max_price",
    "логистика": "fee", "тариф": "fee", "logistics fee": "fee",
}
_REQUIRED = frozenset({"origin", "destination", "min_volume", "fee"})

_FBO_HEADERS = {
    "диапазон объёма от": "min_volume", "диапазон объема от": "min_volume",
    "объём товара": "volume_label", "объем товара": "volume_label",
    "кластер поставки": "origin", "кластер доставки": "destination",
    "для товаров до 300 руб.": "low_fee", "для товаров свыше 300 руб.": "high_fee",
}
_FBO_REQUIRED = frozenset({"origin", "destination", "low_fee", "high_fee"})


def _fallback_min_volume(value: object) -> Decimal:
    text = normalize_text(value).casefold().replace("\xa0", " ")
    match = re.search(r"(?:от\s+)?(\d+(?:[,.]\d+)?)", text)
    if not match:
        raise ValueError
    return parse_decimal(match.group(1))


def _structural_fbo_markers(worksheet):
    """Find FBO section labels accompanied by sibling Unitka section labels."""
    markers = []
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=1, max_row=12, values_only=True), 1
    ):
        normalized = tuple(normalize_header(value) for value in values)
        for fbo_index, name in enumerate(normalized):
            if name != "fbo":
                continue
            fbs_index = next(
                (index for index in range(fbo_index + 1, len(normalized))
                 if normalized[index] == "fbs"),
                None,
            )
            if fbs_index is not None and any(
                name == "базовый тариф" for name in normalized[fbs_index + 1:]
            ):
                markers.append((row_number, fbo_index + 1, fbs_index + 1))
    return tuple(markers)


def _find_fbo_block(worksheet, marker):
    """Return the positional FBO header binding, never worksheet-wide aliases."""
    marker_row, marker_col, next_marker = marker
    for header_row in range(marker_row + 1, min(marker_row + 8, 31)):
        headers = [normalize_header(worksheet.cell(header_row, col).value)
                   for col in range(marker_col, next_marker)]
        binding = {_FBO_HEADERS[name]: marker_col + offset for offset, name in enumerate(headers)
                   if name in _FBO_HEADERS}
        if _FBO_REQUIRED <= binding.keys() and ({"min_volume", "volume_label"} & binding.keys()):
            return header_row, binding
    return None


def _import_fbo(worksheet, header_row, binding, report_context):
    tiers, diagnostics = [], []
    for row_number in range(header_row + 1, worksheet.max_row + 1):
        get = lambda key: worksheet.cell(row_number, binding[key]).value if key in binding else None
        origin, destination = normalize_text(get("origin")), normalize_text(get("destination"))
        if not origin and not destination:
            continue
        if not origin or not destination:
            diagnostics.append(_diag("UNSUPPORTED_UNITKA_TARIFF_LAYOUT", "FBO route is incomplete.", row=row_number)); continue
        try:
            cached = get("min_volume")
            minimum = parse_decimal(cached) if cached is not None and normalize_text(cached) else _fallback_min_volume(get("volume_label"))
            low, high = parse_decimal(get("low_fee")), parse_decimal(get("high_fee"))
            if min(minimum, low, high) < 0: raise ValueError
        except (ValueError, TypeError):
            diagnostics.append(_diag("UNSUPPORTED_UNITKA_TARIFF_LAYOUT", "FBO tariff row cannot be interpreted.", row=row_number)); continue
        tiers.append((origin, destination, minimum, low, high, row_number))
    grouped = defaultdict(list)
    for tier in tiers: grouped[tier[:2]].append(tier)
    records, sources = [], []
    for route in sorted(grouped):
        route_tiers = sorted(grouped[route], key=lambda item: item[2])
        for index, (origin, destination, minimum, low, high, row_number) in enumerate(route_tiers):
            maximum = route_tiers[index + 1][2] if index + 1 < len(route_tiers) else None
            records.extend((TariffRow(origin, destination, minimum, maximum, None, Decimal("300"), low),
                            TariffRow(origin, destination, minimum, maximum, Decimal("300"), None, high)))
            sources.extend((row_number, row_number))
    return ImportResult(tuple(records), tuple(diagnostics), report_context, tuple(sources))


def import_tariffs(data: bytes, report_context: ReportMeta) -> ImportResult[TariffRow]:
    if not data.startswith(b"PK"):
        return ImportResult((), (_diag("TARIFF_SHEET_NOT_FOUND", "Tariffs require an XLSX workbook with a recognizable tariff sheet."),), report_context)
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    fbo_sections = [
        (worksheet, marker)
        for worksheet in workbook.worksheets
        for marker in _structural_fbo_markers(worksheet)
    ]
    fbo_matches = [(worksheet, block) for worksheet, marker in fbo_sections
                   if (block := _find_fbo_block(worksheet, marker)) is not None]
    if fbo_sections:
        if len(fbo_sections) > 1:
            workbook.close()
            return ImportResult((), (_diag("AMBIGUOUS_TARIFF_SHEETS", "Multiple sheets contain an FBO tariff section."),), report_context)
        if not fbo_matches:
            workbook.close()
            return ImportResult((), (_diag(
                "UNSUPPORTED_UNITKA_TARIFF_LAYOUT",
                "The marked FBO tariff section is incomplete or unsupported.",
            ),), report_context)
        worksheet, (header_row, binding) = fbo_matches[0]
        result = _import_fbo(worksheet, header_row, binding, report_context)
        workbook.close()
        return result
    matches = []
    for worksheet in workbook.worksheets:
        for header_row, first in enumerate(worksheet.iter_rows(max_row=30, values_only=True), 1):
            names=tuple(normalize_header(value) for value in first)
            mapped = tuple(_ALIASES.get(name, "") for name in names)
            if _REQUIRED <= set(mapped): matches.append((worksheet.title, mapped, header_row)); break
    if len(matches) != 1:
        workbook.close()
        code = "TARIFF_SHEET_NOT_FOUND" if not matches else "AMBIGUOUS_TARIFF_SHEETS"
        return ImportResult((), (_diag(code, "No tariff sheet matched required columns." if not matches else "Multiple sheets matched tariff columns."),), report_context)
    name, mapped, header_row = matches[0]; worksheet = workbook[name]
    records, sources, diagnostics = [], [], []
    for row_number, values in enumerate(worksheet.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
        row = dict(zip(mapped, values, strict=False))
        if not any(value is not None for value in values): continue
        origin, destination = normalize_text(row.get("origin")), normalize_text(row.get("destination"))
        if not origin or not destination:
            diagnostics.append(_diag("MALFORMED_ROW", "Origin and destination clusters must be nonblank.", row=row_number)); continue
        try:
            min_volume = parse_decimal(row.get("min_volume")); max_volume = parse_decimal(row.get("max_volume"), optional=True)
            min_price = parse_decimal(row.get("min_price"), optional=True); max_price = parse_decimal(row.get("max_price"), optional=True)
            fee = parse_decimal(row.get("fee"))
            if min_volume < 0 or fee < 0 or (min_price is not None and min_price < 0): raise ValueError
        except ValueError:
            diagnostics.append(_diag("INVALID_NUMBER", "Tariff numbers must be finite and non-negative.", row=row_number)); continue
        if max_volume is not None and max_volume < min_volume:
            diagnostics.append(_diag("INVALID_VOLUME_INTERVAL", "Maximum volume must not be below minimum volume.", row=row_number)); continue
        if max_price is not None and (max_price < 0 or min_price is not None and max_price < min_price):
            diagnostics.append(_diag("INVALID_PRICE_INTERVAL", "Maximum price must be non-negative and not below minimum price.", row=row_number)); continue
        records.append(TariffRow(origin, destination, min_volume, max_volume, min_price, max_price, fee)); sources.append(row_number)
    workbook.close()
    return ImportResult(tuple(records), tuple(diagnostics), report_context, tuple(sources))
