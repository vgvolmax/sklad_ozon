"""Small shared row/header utilities for operational importers."""

from dataclasses import dataclass
from io import BytesIO
from math import isfinite
from decimal import Decimal, InvalidOperation

from backend.domain.contracts import ImportDiagnostic
from .csv_adapter import iter_csv_rows
from .normalization import normalize_header


@dataclass(frozen=True, slots=True)
class SourceRows:
    rows: tuple[tuple[int, dict[str, object]], ...]
    diagnostics: tuple[ImportDiagnostic, ...]


def read_source_rows(data: bytes) -> SourceRows:
    if data.startswith(b"PK"):
        from .xlsx import iter_worksheet_rows
        adapted = iter_worksheet_rows(BytesIO(data), 0)
        if not adapted.rows:
            return SourceRows((), adapted.diagnostics + (_diag("MISSING_HEADER", "Worksheet header is missing."),))
        headers = [normalize_header(v) for v in adapted.rows[0].values]
        duplicate = next((h for h in headers if h and headers.count(h) > 1), None)
        if duplicate:
            return SourceRows((), adapted.diagnostics + (_diag("DUPLICATE_HEADER", f"Duplicate normalized header: {duplicate}", field=duplicate),))
        rows = tuple((row.source_row, dict(zip(headers, row.values, strict=False))) for row in adapted.rows[1:])
        return SourceRows(rows, adapted.diagnostics)
    adapted = iter_csv_rows(data)
    return SourceRows(tuple((row.source_row, row.values) for row in adapted.rows), adapted.diagnostics)


def read_xlsx_tables(data: bytes, signature, *, all_sheets=False, scan_rows=30,
                     workbook=None, read_only=False) -> SourceRows:
    """Find logical XLSX headers by signature instead of assuming row one."""
    from openpyxl import load_workbook
    owns_workbook = workbook is None
    if owns_workbook:
        workbook = load_workbook(BytesIO(data), read_only=read_only, data_only=True)
    output, diagnostics = [], []
    for worksheet in workbook.worksheets:
        if read_only and worksheet.calculate_dimension() == "A1:A1":
            worksheet.reset_dimensions()
            diagnostics.append(_diag("WORKSHEET_DIMENSION_REPAIRED", "Worksheet declared range was repaired before row iteration.", severity="warning"))
        header = None
        for number, values in enumerate(worksheet.iter_rows(max_row=scan_rows, values_only=True), 1):
            names = [normalize_header(value) for value in values]
            if signature(names):
                header = number, names
                break
        if header is None:
            continue
        number, names = header
        for row_number, values in enumerate(worksheet.iter_rows(min_row=number + 1, values_only=True), number + 1):
            if any(value is not None and str(value).strip() for value in values):
                output.append((row_number, dict(zip(names, values, strict=False))))
        if not all_sheets:
            break
    if owns_workbook:
        workbook.close()
    if not output and header is None:
        diagnostics.append(_diag("HEADER_ROW_NOT_FOUND", "No worksheet contained the required logical headers."))
    return SourceRows(tuple(output), tuple(diagnostics))


def parse_non_negative_number(value: object) -> float:
    if isinstance(value, bool) or value is None:
        raise ValueError
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        number = float(str(value).replace("\u00a0", "").replace(" ", "").replace(",", "."))
    if not isfinite(number) or number < 0:
        raise ValueError
    return number


def parse_decimal(value: object, *, optional: bool = False) -> Decimal | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        if optional:
            return None
        raise ValueError
    if isinstance(value, bool):
        raise ValueError
    text = str(value).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        result = Decimal(text)
    except (InvalidOperation, ValueError):
        raise ValueError from None
    if not result.is_finite():
        raise ValueError
    return result


def _diag(code: str, message: str, *, row=None, field=None, severity="error"):
    return ImportDiagnostic(severity=severity, code=code, message=message, row=row, field=field)
